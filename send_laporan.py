import pandas as pd
import requests
from typing import List, Dict
import time
import openpyxl
import os
from datetime import datetime, timedelta, timezone

# Konfigurasi
FONNTE_API_KEY = "9eHaFAQMb9pqoGABACCb"
FONNTE_API_URL = "https://api.fonnte.com/send"

# Path ke file Excel (lokal)
EXCEL_FILE = "Laporan.xlsx"

def read_laporan_sheet_by_cell(file_path: str) -> Dict:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['Laporan']
        
        print(f"✓ Sheet 'Laporan' berhasil dibaca")
        
        def get_value(cell, default=''):
            val = cell.value if cell else None
            if val is None:
                return default
            return str(val).strip()

        def format_tahun(tahun):
            try:
                if isinstance(tahun, (int, float)):
                    return str(int(tahun))
                else:
                    tahun_str = str(tahun).split('.')[0]
                    return tahun_str
            except:
                return str(tahun)
              
        def format_date(date_str):
            if not date_str:
                return ''
            try:
                if isinstance(date_str, str):
                    if ' ' in date_str:
                        date_part = date_str.split(' ')[0]
                    else:
                        date_part = date_str
                    
                    if '-' in date_part:
                        parts = date_part.split('-')
                        if len(parts) == 3:
                            return f"{parts[2]}-{parts[1]}-{parts[0]}"
                
                return date_str
            except:
                return date_str
        
        data = {
            'perusahaan': get_value(ws['B3'], 'PT GI REJOSO PASURUAN'),
            'judul_laporan': get_value(ws['C5'], 'LAPORAN RUTIN DGA'),
            'tahun': format_tahun(get_value(ws['C6'], '2025')),
            'tanggal_pengukuran': format_date(get_value(ws['B6'], '')),
            'tanggal_laporan': format_date(get_value(ws['B8'], '')),
            'unit': get_value(ws['G5'], ''),
            'asset': get_value(ws['G6'], ''),
            
            'manufacture': get_value(ws['C11'], ''),
            'type': get_value(ws['C12'], ''),
            'kapasitas': get_value(ws['C13'], ''),
            'serial_number': get_value(ws['C14'], ''),
            'tegangan_primer': get_value(ws['C15'], ''),
            'tegangan_sekunder': get_value(ws['C16'], ''),
            
            'arus_primer': get_value(ws['E11'], ''),
            'arus_sekunder': get_value(ws['E12'], ''),
            'frequency': get_value(ws['E13'], ''),
            'phase': get_value(ws['E14'], ''),
            'temp_rise': get_value(ws['E15'], ''),
            'berat_oli': get_value(ws['E16'], ''),
            
            'hubungan': get_value(ws['G11'], ''),
            'cooling_type': get_value(ws['G12'], ''),
            'tahun_pembuatan': format_tahun(get_value(ws['G13'], '')),
            'impedansi': get_value(ws['G14'], ''),
            'berat_total': get_value(ws['G15'], ''),
            
            'h2_hasil': get_value(ws['D19'], 0),
            'h2_kondisi': get_value(ws['E19'], 0),
            'h2_prealarm': get_value(ws['F19'], 0),
            'h2_alarm': get_value(ws['G19'], 0),
            
            'h2o_hasil': get_value(ws['D20'], 0),
            'h2o_kondisi': get_value(ws['E20'], 0),
            'h2o_prealarm': get_value(ws['F20'], 0),
            'h2o_alarm': get_value(ws['G20'], 0),
            
            'co2_hasil': get_value(ws['D22'], 0),
            'co2_kondisi': get_value(ws['E22'], 0),
            'co2_prealarm': get_value(ws['F22'], 0),
            'co2_alarm': get_value(ws['G22'], 0),
            
            'co_hasil': get_value(ws['D23'], 0),
            'co_kondisi': get_value(ws['E23'], 0),
            'co_prealarm': get_value(ws['F23'], 0),
            'co_alarm': get_value(ws['G23'], 0),
            
            'c2h4_hasil': get_value(ws['D24'], 0),
            'c2h4_kondisi': get_value(ws['E24'], 0),
            'c2h4_prealarm': get_value(ws['F24'], 0),
            'c2h4_alarm': get_value(ws['G24'], 0),
            
            'c2h6_hasil': get_value(ws['D25'], 0),
            'c2h6_kondisi': get_value(ws['E25'], 0),
            'c2h6_prealarm': get_value(ws['F25'], 0),
            'c2h6_alarm': get_value(ws['G25'], 0),
            
            'ch4_hasil': get_value(ws['D26'], 0),
            'ch4_kondisi': get_value(ws['E26'], 0),
            'ch4_prealarm': get_value(ws['F26'], 0),
            'ch4_alarm': get_value(ws['G26'], 0),
            
            'c2h2_hasil': get_value(ws['D27'], 0),
            'c2h2_kondisi': get_value(ws['E27'], 0),
            'c2h2_prealarm': get_value(ws['F27'], 0),
            'c2h2_alarm': get_value(ws['G27'], 0),
            
            'tdcg_hasil': get_value(ws['D28'], 0),
            'tdcg_kondisi': get_value(ws['E28'], 0),
            'tdcg_prealarm': get_value(ws['F28'], 0),
            'tdcg_alarm': get_value(ws['G28'], 0),
            
            'tdcg_analisa': get_value(ws['C32'], 'KONDISI 1'),
            'keygas_analisa': get_value(ws['C33'], 'OVERHEATING OF CELLULOSE'),
            'roger_ratio_analisa': get_value(ws['C34'], 'TIDAK TERDEFINISIKAN'),
            'duval_triangle_analisa': get_value(ws['C35'], 'T3'),
            
            'tdcg_rekomendasi': get_value(ws['C37'], 'Sampling 12 Bulan + Operasi Normal'),
            'keygas_rekomendasi': get_value(ws['C38'], 'Uji PD, Furan, Kadar Air + Cek Umur Isolasi'),
            'roger_ratio_rekomendasi': get_value(ws['C39'], 'GUNAKAN METODE KEYGAS DAN DUVAL TRIANGLE'),
            'duval_triangle_rekomendasi': get_value(ws['C40'], 'LEPAS BEBAN + INSPEKSI INTERNAL TRAFO'),
        }
        
        wb.close()
        return data
        
    except FileNotFoundError:
        print(f"✗ File tidak ditemukan: {file_path}")
        return None
    except Exception as e:
        print(f"✗ Error: {str(e)}")
        return None

        
def get_current_time_wib():
    """Ambil waktu saat ini dalam timezone Indonesia (WIB = UTC+7)"""
    tz_wib = timezone(timedelta(hours=7))
    now = datetime.now(tz_wib)
    return now.strftime('%d-%m-%Y %H:%M')

def format_laporan_to_whatsapp_single(data: Dict) -> str:
    message = f"*🏢 {data['perusahaan']}*\n"
    message += f"*📊 {data['judul_laporan']} - {data['tahun']}*\n"
    message += "=" * 50 + "\n\n"
    
    message += "*📅 INFORMASI DASAR*\n"
    message += f"Tanggal Pengukuran: {data['tanggal_pengukuran']}\n"
    message += f"Tanggal Laporan: {data['tanggal_laporan']}\n"
    message += f"Unit: {data['unit']}\n"
    message += f"Asset: {data['asset']}\n\n"
    
    message += "*⚙️ SPESIFIKASI TRAFO*\n"
    message += f"Manufacture: {data['manufacture']}\n"
    message += f"Kapasitas: {data['kapasitas']}\n"
    message += f"Serial Number: {data['serial_number']}\n"
    message += f"Tegangan Primer: {data['tegangan_primer']}\n"
    message += f"Tegangan Sekunder: {data['tegangan_sekunder']}\n"
    
    if data['arus_primer']:
        message += f"Arus Primer: {data['arus_primer']}\n"
    if data['arus_sekunder']:
        message += f"Arus Sekunder: {data['arus_sekunder']}\n"
    
    message += f"Frequency: {data['frequency']}\n"
    message += f"Phase: {data['phase']}\n"
    
    if data['temp_rise']:
        message += f"Temp Rise: {data['temp_rise']}\n"
    
    message += f"Kapasitas Oil: {data['berat_oli']}\n"
    message += f"Konfigurasi: {data['hubungan']}\n"
    message += f"Cooling: {data['cooling_type']}\n"
    message += f"Tahun Pembuatan: {data['tahun_pembuatan']}\n"
    
    message += "\n"
    
    message += "*🔬 CONTENT ANALYSIS*\n\n"
    
    gas_data = [
        ("Hydrogen", "H2", data['h2_hasil'], data['h2_kondisi'], data['h2_prealarm'], data['h2_alarm']),
        ("Air", "H2O", data['h2o_hasil'], data['h2o_kondisi'], data['h2o_prealarm'], data['h2o_alarm']),
        ("Karbon Dioksida", "CO2", data['co2_hasil'], data['co2_kondisi'], data['co2_prealarm'], data['co2_alarm']),
        ("Karbon Monoksida", "CO", data['co_hasil'], data['co_kondisi'], data['co_prealarm'], data['co_alarm']),
        ("Etilen", "C2H4", data['c2h4_hasil'], data['c2h4_kondisi'], data['c2h4_prealarm'], data['c2h4_alarm']),
        ("Etana", "C2H6", data['c2h6_hasil'], data['c2h6_kondisi'], data['c2h6_prealarm'], data['c2h6_alarm']),
        ("Metana", "CH4", data['ch4_hasil'], data['ch4_kondisi'], data['ch4_prealarm'], data['ch4_alarm']),
        ("Asetilen", "C2H2", data['c2h2_hasil'], data['c2h2_kondisi'], data['c2h2_prealarm'], data['c2h2_alarm']),
        ("Total Gas", "TDCG", data['tdcg_hasil'], data['tdcg_kondisi'], data['tdcg_prealarm'], data['tdcg_alarm']),
    ]
    
    for nama_gas, simbol, hasil, kondisi, prealarm, alarm in gas_data:
        message += f"{nama_gas} ({simbol})\n"
        message += f"  • Hasil Pengujian: {hasil}  | Kondisi 2: {kondisi}  | Pre-Alarm: {prealarm}  | Alarm: {alarm}\n"
        message += "\n"
    
    message += "*🔍 HASIL ANALISA*\n\n"
    message += f"TDCG:\n{data['tdcg_analisa']}\n\n"
    message += f"KEY GAS:\n{data['keygas_analisa']}\n\n"
    message += f"ROGER RATIO:\n{data['roger_ratio_analisa']}\n\n"
    message += f"DUVAL TRIANGLE:\n{data['duval_triangle_analisa']}\n\n"
    
    message += "*📌 REKOMENDASI*\n\n"
    message += f"TDCG:\n{data['tdcg_rekomendasi']}\n\n"
    message += f"KEY GAS:\n{data['keygas_rekomendasi']}\n\n"
    message += f"ROGER RATIO:\n{data['roger_ratio_rekomendasi']}\n\n"
    message += f"DUVAL TRIANGLE:\n{data['duval_triangle_rekomendasi']}\n\n"
    
    message += "=" * 50 + "\n"
    message += f"Generated: {get_current_time_wib()}\n"
    
    return message

def send_whatsapp_message(phone_number: str, message: str, is_group: bool = False) -> bool:
    try:
        headers = {
            "Authorization": FONNTE_API_KEY,
            "Content-Type": "application/json"
        }
        
        if is_group:
            payload = {
                "target": phone_number,
                "message": message
            }
        else:
            payload = {
                "target": phone_number,
                "message": message,
                "countryCode": "62"
            }
        
        response = requests.post(
            FONNTE_API_URL,
            headers=headers,
            json=payload,
            timeout=10
        )
        
        if response.status_code == 200:
            result = response.json()
            if result.get("status"):
                device_type = "grup" if is_group else "personal"
                print(f"  ✓ Pesan berhasil dikirim ke {device_type}: {phone_number}")
                return True
            else:
                print(f"  ✗ Gagal: {result.get('reason', 'Unknown error')}")
                return False
        else:
            print(f"  ✗ Error API: Status {response.status_code}")
            return False
            
    except Exception as e:
        print(f"  ✗ Error: {str(e)}")
        return False

# ===== MAIN =====

if __name__ == "__main__":
    print("\n" + "="*70)
    print("📋 PROGRAM KIRIM LAPORAN DGA KE WHATSAPP")
    print(f"⏰ Waktu: {get_current_time_wib()}")
    print("="*70 + "\n")
    
    # Cek file ada
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ File tidak ditemukan: {EXCEL_FILE}")
        exit(1)
    
    print("1️⃣  Membaca file Excel...")
    data = read_laporan_sheet_by_cell(EXCEL_FILE)
    
    if data is None:
        print("\n❌ Program dihentikan!")
        exit(1)
    
    print("2️⃣  Memformat laporan...")
    message = format_laporan_to_whatsapp_single(data)
    print(f"  ✓ Laporan siap: {len(message)} karakter\n")
    
    phone_numbers = [
        {"number": "6282157905834", "type": "personal"},
        {"number": "120363426691870718@g.us", "type": "group"},
    ]
    
    print(f"3️⃣  Mengirim ke {len(phone_numbers)} nomor/grup WhatsApp...")
    print("-" * 70)
    
    success_total = 0
    for item in phone_numbers:
        is_group = item.get('type') == 'group'
        phone_number = item.get('number')
        device_type = "Grup" if is_group else "Personal"
        
        print(f"\n📱 [{device_type}] Mengirim ke: {phone_number}")
        if send_whatsapp_message(phone_number, message, is_group):
            success_total += 1
        time.sleep(1)
    
    print("\n" + "="*70)
    if success_total == len(phone_numbers):
        print(f"✅ BERHASIL! {success_total}/{len(phone_numbers)} pesan terkirim")
    else:
        print(f"⚠️  {success_total}/{len(phone_numbers)} pesan terkirim")
    print("="*70 + "\n")
